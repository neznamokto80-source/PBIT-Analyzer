from datetime import datetime
import logging

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    SHEETS_ORDER = [
        ("Справочник таблиц", "tables_ref"),
        ("Столбцы", "columns"),
        ("Источники", "sources"),
        ("Связи", "relationships"),
        ("Меры", "measures"),
        ("Листы", "sheets"),
        ("Визуальные элементы", "visuals"),
        ("Линейдж", "lineage"),
        ("Lineage_Upstream", "lineage_upstream"),
        ("Lineage_Downstream", "lineage_downstream"),
        ("Условное форматирование", "conditional"),
        ("Статистика", "stats"),
    ]

    def __init__(self, output_path: str, dfs: dict, source_file: str):
        self.output_path = output_path
        self.dfs = dfs
        self.source_file = source_file

    def generate(self):
        with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            for sheet_name, df_key in self.SHEETS_ORDER:
                df = self.dfs.get(df_key)
                if df is not None and not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    cols = self._get_columns(df_key)
                    pd.DataFrame(columns=cols).to_excel(writer, sheet_name=sheet_name, index=False)
                self._format_sheet(writer, sheet_name)

                # Для "Визуальные элементы" — объединение ячеек
                if df_key == "visuals" and df is not None and not df.empty:
                    self._merge_visuals_cells(writer, sheet_name, df)

        logger.info("Excel-файл сохранён: %s", self.output_path)

    def _merge_visuals_cells(self, writer, sheet_name: str, df: pd.DataFrame):
        """Объединяет ячейки для общих атрибутов визуала (Имя листа, ID, Тип, Расположение).
        Логика: для каждого визуала находим все его строки, объединяем
        столбцы 1,2,3,10 (Имя листа, ID визуала, Тип визуала, Расположение)."""
        try:
            wb = writer.book
            ws = wb[sheet_name]
            if ws.max_row < 2:
                return

            # Колонки для объединения (1-based, соответствуют порядку в df)
            # Имя листа=1, ID визуала=2, Тип визуала=3, Расположение=10
            merge_cols = [1, 2, 3, 10]

            # Проходим по строкам и находим группы по ID визуала
            # ID визуала в колонке 2. Первая строка — заголовок (row 1), данные с row 2.
            visual_id_col = 2  # колонка B
            i = 2  # начинаем с первой строки данных
            while i <= ws.max_row:
                # Находим визуал (если ячейка None, идём вверх)
                cell_val = ws.cell(row=i, column=visual_id_col).value
                if cell_val is None or cell_val == '':
                    i += 1
                    continue

                visual_id = cell_val
                # Находим последнюю строку этого визуала
                end = i
                j = i + 1
                while j <= ws.max_row:
                    next_val = ws.cell(row=j, column=visual_id_col).value
                    if next_val is not None and next_val != '' and next_val != visual_id:
                        break
                    end = j
                    j += 1

                if end > i:
                    # Объединяем
                    for col_num in merge_cols:
                        if col_num <= ws.max_column:
                            ws.merge_cells(start_row=i, start_column=col_num,
                                           end_row=end, end_column=col_num)
                            # Выравнивание для объединённых ячеек
                            for r in range(i, end + 1):
                                cell = ws.cell(row=r, column=col_num)
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                i = end + 1

        except Exception as e:
            logger.warning("Предупреждение при объединении ячеек: %s", e)

    def _get_columns(self, df_key: str):
        col_map = {
            "tables_ref": ["№ таблицы", "Имя таблицы", "Описание"],
            "columns": ["№ столбца", "№ таблицы", "Таблица", "Столбец",
                         "Описание", "Тип данных", "Тип значения", "Формат",
                         "Сортировка по столбцу", "Агрегация по умолчанию",
                         "Формула (если вычисляемый)", "Зависимости"],
            "sources": ["№ таблицы", "Таблица", "Тип источника", "Как создается", "Зависимости"],
            "relationships": ["Название связи", "Откуда", "Куда", "Состояние"],
            "measures": ["№ меры", "№ таблицы", "Таблица", "Мера", "Описание", "Формула (DAX)",
                          "Формат", "Папка отображения", "Зависимости"],
            "sheets": ["№ листа", "Имя листа", "ID листа", "Видимость", "Количество визуалов", "Размеры"],
            "visuals": [
                "Имя листа", "ID визуала", "Тип визуала",
                "Роль", "Порядок", "Поле", "Отображаемое имя",
                "№ таблицы", "№ столбца/меры", "Расположение",
                "FilterTargetEntity", "FilterTargetProperty", "FilterExpressionShort",
            ],
            "conditional": [
                "Имя листа", "ID визуала", "Поле", "Элемент", "Правило", "Цвет",
            ],
            "lineage": [
                "Тип источника",
                "ID источника",
                "Имя источника",
                "Тип приемника",
                "ID приемника",
                "Имя приемника",
                "Тип связи",
                "Контекст",
            ],
            "lineage_upstream": [
                "ID фокуса",
                "Тип фокуса",
                "Имя фокуса",
                "Тип источника",
                "ID источника",
                "Имя источника",
                "Тип приемника",
                "ID приемника",
                "Имя приемника",
                "Тип связи",
                "Контекст",
            ],
            "lineage_downstream": [
                "ID фокуса",
                "Тип фокуса",
                "Имя фокуса",
                "Тип источника",
                "ID источника",
                "Имя источника",
                "Тип приемника",
                "ID приемника",
                "Имя приемника",
                "Тип связи",
                "Контекст",
            ],
            "stats": ["Параметр", "Значение"],
        }
        return col_map.get(df_key, [])

    def _format_sheet(self, writer, sheet_name: str):
        wb = writer.book
        ws = wb[sheet_name]
        if ws.max_row < 1:
            return
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = DATA_ALIGNMENT
                cell.border = THIN_BORDER
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 10
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    lines = str(val).split('\n')
                    longest = max(len(line) for line in lines) if lines else 0
                    max_len = max(max_len, longest)
            ws.column_dimensions[letter].width = min(max_len + 2, 80)
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    @staticmethod
    def build_stats(dfs: dict, source_file: str):
        stats = {
            "Всего таблиц": len(dfs.get("tables_ref", [])),
            "Всего столбцов": len(dfs.get("columns", [])),
            "Всего источников": len(dfs.get("sources", [])),
            "Всего мер": len(dfs.get("measures", [])),
            "Всего связей": len(dfs.get("relationships", [])),
            "Всего листов": len(dfs.get("sheets", [])),
            "Всего визуальных элементов": len(dfs.get("visuals", [])),
            "Всего связей lineage": len(dfs.get("lineage", [])),
            "Всего связей lineage upstream": len(dfs.get("lineage_upstream", [])),
            "Всего связей lineage downstream": len(dfs.get("lineage_downstream", [])),
            "Всего правил условного форматирования": len(dfs.get("conditional", [])),
            "Файл источника": source_file,
            "Дата обработки": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        return pd.DataFrame(list(stats.items()), columns=["Параметр", "Значение"])


# ============================================================
# 6. Utility
# ============================================================

